"""
Генератор Excel-шаблона с предзаполненными данными из распознанной схемы.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BH = "1F4E79"; BL = "BDD7EE"; YL = "FFF2CC"; GR = "E2EFDA"; GY = "F2F2F2"


def _tb():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, r, c, txt, bg=BH, fg="FFFFFF", bold=True, sz=11, r2=None, c2=None):
    if r2 and c2:
        ws.merge_cells(start_row=r, start_column=c, end_row=r2, end_column=c2)
    cl = ws.cell(row=r, column=c); cl.value = txt
    cl.font = Font(bold=bold, color=fg, size=sz)
    cl.fill = PatternFill("solid", fgColor=bg)
    cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _inp(ws, r, c, v=None, fmt=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.fill = PatternFill("solid", fgColor=YL)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    if fmt: cl.number_format = fmt
    return cl


def _res(ws, r, c, f, fmt="0.0000000000"):
    cl = ws.cell(row=r, column=c, value=f)
    cl.fill = PatternFill("solid", fgColor=GR)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.number_format = fmt
    return cl


def generate_excel(scheme_data: dict, cos_phi: float = 0.90,
                    k2f: float = 1.10, kk: float = 0.99) -> bytes:
    """
    Создаёт Excel с предзаполненными данными из схемы.
    Возвращает bytes для отправки в Telegram.
    """
    wb = Workbook()

    lines = scheme_data.get("lines", [])
    trs   = scheme_data.get("transformers", [])

    # ── Лист 1: Параметры (ввод / проверка) ──
    ws1 = wb.active; ws1.title = "1. Параметры сети"
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [4, 30, 12, 14, 14, 14, 14, 16, 14, 20]):
        ws1.column_dimensions[col].width = w

    _hdr(ws1, 1, 1, "РАСЧЁТ ПОТЕРЬ ЭЛЕКТРОЭНЕРГИИ — ПАРАМЕТРЫ СЕТИ", r2=1, c2=9, sz=13)
    ws1.row_dimensions[1].height = 28
    _hdr(ws1, 2, 1,
         "Приказ Минэнерго РФ № 326 от 30.12.2008  |  Данные получены из схемы автоматически — проверь и исправь если нужно",
         bg=BL, fg="1F4E79", bold=False, r2=2, c2=9, sz=10)

    # Общие коэффициенты
    _hdr(ws1, 4, 1, "А.  ОБЩИЕ КОЭФФИЦИЕНТЫ", r2=4, c2=9)
    _hdr(ws1, 5, 2, "Параметр", bg=BL, fg=BH); _hdr(ws1, 5, 3, "Значение", bg=BL, fg=BH)
    _hdr(ws1, 5, 4, "Описание", bg=BL, fg=BH, r2=5, c2=9)
    coeffs = [
        (6, "K²ф — коэф. формы графика нагрузки", k2f,    "уже в квадрате"),
        (7, "k_k — коэф. конфигурации графиков",  kk,     "принимается 0.99"),
        (8, "cos φ средневзвешенный",              cos_phi,"по данным учёта"),
    ]
    for r, lab, val, desc in coeffs:
        ws1.cell(row=r, column=2, value=lab).alignment = Alignment(indent=1, vertical="center")
        _inp(ws1, r, 3, val)
        c = ws1.cell(row=r, column=4, value=desc)
        c.font = Font(italic=True, color="606060", size=9)
        c.alignment = Alignment(indent=1); ws1.merge_cells(f"D{r}:I{r}")
    for r in range(5, 9):
        for ci in range(2, 5): ws1.cell(row=r, column=ci).border = _tb()

    # Линии
    _hdr(ws1, 10, 1, "Б.  ЛИНИИ — ΔW = C × W²а / T", r2=10, c2=9)
    for ci, h in enumerate(["Наименование","Тип","Марка","r₀, Ом/км","L, км","U, кВ","Группа (цвет/парал.)","Счётчик","Примечание"], 2):
        c = ws1.cell(row=11, column=ci, value=h)
        c.font = Font(bold=True, color=BH, size=9)
        c.fill = PatternFill("solid", fgColor=BL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[11].height = 28

    for ri, line in enumerate(lines[:15], 12):
        ws1.cell(row=ri, column=2, value=line.get("name", "")).alignment = Alignment(indent=1)
        _inp(ws1, ri, 3, line.get("type", ""))
        _inp(ws1, ri, 4, line.get("mark", ""))
        _inp(ws1, ri, 5, line.get("r0"))
        _inp(ws1, ri, 6, line.get("length"))
        _inp(ws1, ri, 7, line.get("voltage", 10))
        _inp(ws1, ri, 8, line.get("color_group") or "")   # группа параллельных
        _inp(ws1, ri, 9, "")
        _inp(ws1, ri, 10, line.get("note", ""))
    # Пустые строки
    for ri in range(12 + len(lines), 22):
        for ci in range(2, 10): _inp(ws1, ri, ci)

    for r in range(11, 22):
        for ci in range(2, 10): ws1.cell(row=r, column=ci).border = _tb()

    # Трансформаторы
    _hdr(ws1, 23, 1, "В.  ТРАНСФОРМАТОРЫ — ΔW = A × T + B × W²а / T", r2=23, c2=9)
    for ci, h in enumerate(["Наименование","Марка","ΔP₀, кВт","ΔPк, кВт","Sн, кВА","n, шт","Счётчик","Примечание"], 2):
        c = ws1.cell(row=24, column=ci, value=h)
        c.font = Font(bold=True, color=BH, size=9)
        c.fill = PatternFill("solid", fgColor=BL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, tr in enumerate(trs[:8], 25):
        ws1.cell(row=ri, column=2, value=tr.get("name", "")).alignment = Alignment(indent=1)
        _inp(ws1, ri, 3, tr.get("mark", ""))
        _inp(ws1, ri, 4, tr.get("P0"))
        _inp(ws1, ri, 5, tr.get("Pk"))
        _inp(ws1, ri, 6, tr.get("Sn"))
        _inp(ws1, ri, 7, tr.get("n", 1))
        _inp(ws1, ri, 8, "")
        _inp(ws1, ri, 9, tr.get("note", ""))
    for ri in range(25 + len(trs), 33):
        for ci in range(2, 10): _inp(ws1, ri, ci)

    for r in range(24, 33):
        for ci in range(2, 10): ws1.cell(row=r, column=ci).border = _tb()

    note = ws1.cell(row=34, column=2,
        value="⚠  Проверь все жёлтые ячейки. Исправь ошибки распознавания. Загрузи файл обратно в бот.")
    note.font = Font(bold=True, color="C00000", size=10)
    ws1.merge_cells("B34:I34")

    # ── Лист 2: Расчёт констант линий ──
    ws2 = wb.create_sheet("2. Константы линий")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [4, 28, 10, 12, 12, 12, 16, 34]):
        ws2.column_dimensions[col].width = w

    _hdr(ws2, 1, 1, "КОНСТАНТА C — ΔW_лин = C × W²а / T", r2=1, c2=8, sz=12)
    _hdr(ws2, 2, 1, "C = K²ф × k_k × r₀ × L / (U² × cos²φ × 10³)",
         bg=BL, fg="1F4E79", bold=True, r2=2, c2=8, sz=11)

    I = "'1. Параметры сети'"
    ws2.row_dimensions[4].height = 30
    for ci, h in enumerate(["№","Наименование","r₀·L (Ом)","U² (кВ²)","cos²φ","K²ф·k_k","КОНСТАНТА C","ФОРМУЛА"], 1):
        c = ws2.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color=BH, size=9)
        c.fill = PatternFill("solid", fgColor=BL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in range(10):
        r = i + 5; s = 12 + i
        ws2.cell(row=r, column=1, value=i+1).alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GY)
        c = ws2.cell(row=r, column=2, value=f"={I}!B{s}")
        c.fill = PatternFill("solid", fgColor=GY); c.alignment = Alignment(indent=1)
        _res(ws2, r, 3, f"=IF(AND({I}!E{s}<>\"\",{I}!F{s}<>\"\"),{I}!E{s}*{I}!F{s},\"\")", "0.0000")
        _res(ws2, r, 4, f"=IF({I}!G{s}<>\"\",{I}!G{s}^2,\"\")", "0.00")
        _res(ws2, r, 5, f"=IF({I}!E{s}<>\"\",{I}!$C$8^2,\"\")", "0.0000")
        _res(ws2, r, 6, f"=IF({I}!E{s}<>\"\",{I}!$C$6*{I}!$C$7,\"\")", "0.0000")
        _res(ws2, r, 7, f"=IF(AND(D{r}<>\"\",E{r}<>\"\"),F{r}*C{r}/(D{r}*E{r}*1000),\"\")", "0.0000000000", )
        ws2.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFE699")
        f_txt = f'=IF(G{r}<>"","ΔW = "&TEXT(G{r},"0.000E+00")&" × W²а / T","")'
        c = ws2.cell(row=r, column=8, value=f_txt)
        c.fill = PatternFill("solid", fgColor="DDEBF7"); c.alignment = Alignment(indent=1)
        for ci in range(1, 9): ws2.cell(row=r, column=ci).border = _tb()

    ri = 15
    _hdr(ws2, ri, 1, "ИТОГО C", r2=ri, c2=6)
    _res(ws2, ri, 7, "=SUM(G5:G14)", "0.0000000000")
    ws2.cell(row=ri, column=7).fill = PatternFill("solid", fgColor="FFD700")
    c = ws2.cell(row=ri, column=8, value=f'="ΔW_лин = "&TEXT(G{ri},"0.000E+00")&" × W²а / T"')
    c.fill = PatternFill("solid", fgColor="FFD700"); c.font = Font(bold=True, size=11)
    c.alignment = Alignment(indent=1)
    for ci in range(1, 9): ws2.cell(row=ri, column=ci).border = _tb()

    # ── Лист 3: Константы трансформаторов ──
    ws3 = wb.create_sheet("3. Константы трансформаторов")
    ws3.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [4, 26, 12, 12, 12, 10, 12, 14, 14, 34]):
        ws3.column_dimensions[col].width = w

    _hdr(ws3, 1, 1, "КОНСТАНТЫ A и B — ΔW_тр = A × T + B × W²а / T", r2=1, c2=10, sz=12)
    _hdr(ws3, 2, 1, "A = ΔP₀ × n  (постоянные потери, кВт)", bg=BL, fg="1F4E79", bold=True, r2=2, c2=10, sz=11)
    _hdr(ws3, 3, 1, "B = ΔPк × n / (S²н × cos²φ)  (нагрузочная составляющая)", bg=BL, fg="1F4E79", bold=False, r2=3, c2=10, sz=10)

    ws3.row_dimensions[5].height = 30
    for ci, h in enumerate(["№","Наименование","ΔP₀","ΔPк","Sн","n","cos²φ","КОНСТ. A","КОНСТ. B","ФОРМУЛА"], 1):
        c = ws3.cell(row=5, column=ci, value=h)
        c.font = Font(bold=True, color=BH, size=9)
        c.fill = PatternFill("solid", fgColor=BL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in range(8):
        r = i + 6; s = 25 + i
        ws3.cell(row=r, column=1, value=i+1).alignment = Alignment(horizontal="center")
        ws3.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GY)
        c = ws3.cell(row=r, column=2, value=f"={I}!B{s}")
        c.fill = PatternFill("solid", fgColor=GY); c.alignment = Alignment(indent=1)
        for ci, sc in [(3,"D"),(4,"E"),(5,"F"),(6,"G")]:
            c = ws3.cell(row=r, column=ci, value=f"={I}!{sc}{s}")
            c.fill = PatternFill("solid", fgColor=GY); c.alignment = Alignment(horizontal="center")
        _res(ws3, r, 7, f"=IF({I}!D{s}<>\"\",{I}!$C$8^2,\"\")", "0.0000")
        _res(ws3, r, 8, f"=IF(AND(C{r}<>\"\",F{r}<>\"\"),C{r}*F{r},\"\")", "0.000")
        ws3.cell(row=r, column=8).fill = PatternFill("solid", fgColor="FFE699")
        _res(ws3, r, 9, f"=IF(AND(D{r}<>\"\",E{r}<>\"\",F{r}<>\"\"),D{r}*F{r}/(E{r}^2*G{r}),\"\")", "0.0000000000")
        ws3.cell(row=r, column=9).fill = PatternFill("solid", fgColor="FFE699")
        f_txt = f'=IF(AND(H{r}<>"",I{r}<>""),"ΔW = "&TEXT(H{r},"0.000")&"×T + "&TEXT(I{r},"0.000E+00")&"×W²а/T","")'
        c = ws3.cell(row=r, column=10, value=f_txt)
        c.fill = PatternFill("solid", fgColor="DDEBF7"); c.alignment = Alignment(indent=1)
        for ci in range(1, 11): ws3.cell(row=r, column=ci).border = _tb()

    ri = 14
    _hdr(ws3, ri, 1, "ИТОГО A и B", r2=ri, c2=7)
    _res(ws3, ri, 8, "=SUM(H6:H13)", "0.000")
    ws3.cell(row=ri, column=8).fill = PatternFill("solid", fgColor="FFD700")
    _res(ws3, ri, 9, "=SUM(I6:I13)", "0.0000000000")
    ws3.cell(row=ri, column=9).fill = PatternFill("solid", fgColor="FFD700")
    c = ws3.cell(row=ri, column=10,
        value=f'="ΔW_тр = "&TEXT(H{ri},"0.000")&"×T + "&TEXT(I{ri},"0.000E+00")&"×W²а/T"')
    c.fill = PatternFill("solid", fgColor="FFD700"); c.font = Font(bold=True, size=11)
    c.alignment = Alignment(indent=1)
    for ci in range(1, 11): ws3.cell(row=ri, column=ci).border = _tb()

    # Сохраняем в bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
